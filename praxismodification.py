

import openpyxl,logging, os
logging.basicConfig(level=logging.WARNING, format=' %(asctime)s - %(levelname)s- %(message)s')
os.chdir("M:\Assessment (New)\\CAEP\\Future Accreditation\\Praxis Scores Last Three Years")

wb = openpyxl.load_workbook("CompletersandPraxis.xlsx")

sheet = wb.get_sheet_by_name(name='etscopy')
newsheet = wb.get_sheet_by_name(name='tests')
newsheetrow =2
newsheetcolumn = 2 

for row in range(2, 3118):
	candidateid = sheet['E' + str(row)].value
	nextcandidateid = sheet['E'+str(row +1)].value
	testname = sheet['W'+ str(row)].value
	nexttestname = sheet['W'+ str(row +1)].value
	testscore = sheet['Y' + str(row)].value
	name = sheet['D' + str(row)].value
	if (candidateid==nextcandidateid):
		newsheetcolumn += 1
		newsheet.cell(row=newsheetrow, column=newsheetcolumn, value=testname)
		newsheetcolumn += 1
		newsheet.cell(row=newsheetrow, column=newsheetcolumn, value=testscore)
		logging.warning("{}".format(newsheetcolumn))
	else:
		newsheetcolumn+=1
		newsheet.cell(row=newsheetrow, column=newsheetcolumn, value=testname)
		newsheetcolumn +=1
		newsheet.cell(row=newsheetrow, column=newsheetcolumn, value=testscore)
		newsheetcolumn = 2
		newsheet.cell(row=newsheetrow, column=newsheetcolumn, value=name)
		newsheetrow+=1
		

wb.save("diditwork.xlsx")