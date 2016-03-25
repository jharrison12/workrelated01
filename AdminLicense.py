#File created on 3/25/2016
#Jonathan Harrison
#File that opens a workbook and transposes nine rows and three columns across one row for various students


import openpyxl
import webbrowser, pyperclip, requests, bs4, openpyxl, os, logging, requests, re, getpass

from openpyxl.utils import column_index_from_string
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait # available since 2.4.0
from selenium.webdriver.support import expected_conditions as EC # available since 2.26.0

os.chdir('M:\Working Folder')

wb = openpyxl.load_workbook('admin.xlsx')
sheet = wb.get_sheet_by_name('Original')
scores = wb.get_sheet_by_name("Scores")

categories = ["Cat IIA Teaching and Learning",
              "Cat III Managing Organizational Systems and Safety",
              "Cat IV Collaborating with Key Stakeholders",
              "Cat V Ethics and Integrity",
              "Cat VI The Educational System",
              "Cat IB Vision and Goals (constructed response)",
              "Cat IIB Teaching and Learning (constructed response)",
              "Cat IIA Teaching and Learning",
              "Cat III Managing Organizational Systems and Safety",
              "Cat IV Collaborating with Key Stakeholders",
              "Cat V Ethics and Integrity",
              "Cat VI The Educational System",
              "Cat IB Vision and Goals (constructed response)",
              "Cat IIB Teaching and Learning (constructed response)"]

name = []
namecolumn = column_index_from_string('F')

#TODO Take a row of four cells and print those cells on a new sheet
#Matching with the correct name.  Once you have gone through those cells
#Iterate to the next row and do the same
#
initialcol = 2
initialrow = 1


for scoresnamerow in range(1, scores.max_row):
    for rows in range(1, sheet.max_row):
        if scores.cell(row=scoresnamerow, column=column_index_from_string("A")).value == sheet.cell(row=rows, column=namecolumn).value:
            initialrow += 1
            initialcol = 2
            for cat in range(9):
                for testscore in range(3):
                    scores.cell(row=initialrow, column=initialcol+testscore).value = \
                    sheet.cell(row=(rows + cat), column=(column_index_from_string('T') + testscore)).value
                initialcol += 3
        else:
            continue
        break



wb.save('NEWADMIN.xlsx')